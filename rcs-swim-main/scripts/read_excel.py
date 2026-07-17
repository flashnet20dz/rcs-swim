"""قراءة بيانات ملف Excel وفهم بنية كل ورقة"""
import openpyxl
from openpyxl import load_workbook

wb = load_workbook("upload/منظومة_اشتراكات_RCS_v2.xlsx", data_only=True, read_only=True)

for sheet_name in ["بيانات", "قائمة_التأمين", "حقوق_المركب"]:
    print(f"\n{'='*70}")
    print(f"الورقة: {sheet_name}")
    print('='*70)
    ws = wb[sheet_name]
    for i, row in enumerate(ws.iter_rows(max_row=8, values_only=True)):
        print(f"الصف {i+1}: {row}")
    print()

wb.close()
