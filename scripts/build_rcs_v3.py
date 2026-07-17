#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
إعادة بناء منظومة اشتراكات RCS v3
- ملء كل سطور البيانات بالمعادلات
- المبلغ الإجمالي = رسوم الاشتراك + مصاريف التأمين
- OPOW / DJS = 300 دج اشتراك + 500 دج تأمين
- تحويل حقول الخيارات إلى خانات يضع فيها المستخدم علامة X أو •
"""

import os
from copy import copy
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule

SRC = "/home/z/my-project/upload/منظومة_اشتراكات_RCS_v2.xlsx"
DST = "/home/z/my-project/download/منظومة_اشتراكات_RCS_v3.xlsx"

# ───────────────── ألوان وخطوط (من الملف الأصلي) ─────────────────
COLOR_TITLE_BG   = "FF0D1B2A"
COLOR_HEADER_BG  = "FF1B3A6B"
COLOR_DATE_BG    = "FF0F766E"
COLOR_BLOOD_BG   = "FF6D28D9"
COLOR_PAY_BG     = "FF1D4ED8"
COLOR_MONEY_BG   = "FF0F766E"
COLOR_TOTAL_BG   = "FFD97706"
COLOR_RENEW_BG   = "FF15803D"
COLOR_TIME_BG    = "FF6D28D9"
COLOR_MONTHS_BG  = "FF1F2937"
COLOR_FOOTER_BG  = "FF111827"

# ألوان الخلفية للخانات
COLOR_X_BG       = "FFFFFBEB"   # خانات علامة X (أصفر فاتح)
COLOR_X_HEADER   = "FFB45309"   # رأس خانات X (برتقالي غامق)
COLOR_AUTO_BG    = "FFF0FDFB"   # الخلايا التلقائية
COLOR_INPUT_BG   = "FFFFFFFF"   # خلايا الإدخال
COLOR_GROUP_BG   = "FFFEF3C7"   # خلفية مجموعة الدفع

FONT_NAME   = "FreeSans"
FONT_NAME_NUM = "Calibri"

# حدود
thin = Side(style="thin", color="FF9CA3AF")
medium = Side(style="medium", color="FF1B3A6B")
BORDER_ALL  = Border(left=thin, right=thin, top=thin, bottom=thin)
BORDER_BOX  = Border(left=medium, right=medium, top=medium, bottom=medium)

# ───────────────── تعريف بنية الأعمدة الجديدة ─────────────────
# (col_letter, header, kind, width, bg_color, x_options)
# kind: 'auto' | 'input' | 'x' | 'title' | 'group'
# لأعمدة X: x_options = قائمة الخيارات المرتبة بنفس ترتيب الأعمدة

COLUMNS = [
    # A
    ("A", "رقم\nالملف", "auto", 9, COLOR_HEADER_BG, None),
    ("B", "اللقب", "input", 18, COLOR_HEADER_BG, None),
    ("C", "الاسم", "input", 16, COLOR_HEADER_BG, None),
    ("D", "تاريخ\nالميلاد", "input", 12, COLOR_DATE_BG, None),
    # الجنس
    ("E", "ذكر", "x", 5, COLOR_X_HEADER, "ذكر"),
    ("F", "أنثى", "x", 5, COLOR_X_HEADER, "أنثى"),
    ("G", "الجنس", "auto", 8, COLOR_DATE_BG, None),
    ("H", "العمر", "auto", 6, COLOR_DATE_BG, None),
    # فصيلة الدم (8 خيارات)
    ("I", "A+", "x", 4, COLOR_X_HEADER, "A+"),
    ("J", "A-", "x", 4, COLOR_X_HEADER, "A-"),
    ("K", "B+", "x", 4, COLOR_X_HEADER, "B+"),
    ("L", "B-", "x", 4, COLOR_X_HEADER, "B-"),
    ("M", "O+", "x", 4, COLOR_X_HEADER, "O+"),
    ("N", "O-", "x", 4, COLOR_X_HEADER, "O-"),
    ("O", "AB+", "x", 4, COLOR_X_HEADER, "AB+"),
    ("P", "AB-", "x", 4, COLOR_X_HEADER, "AB-"),
    ("Q", "فصيلة\nالدم", "auto", 8, COLOR_BLOOD_BG, None),
    # نوع الاشتراك (6 خيارات)
    ("R", "/", "x", 4, COLOR_X_HEADER, "/"),
    ("S", "OPOW", "x", 5, COLOR_X_HEADER, "OPOW"),
    ("T", "DJS", "x", 5, COLOR_X_HEADER, "DJS"),
    ("U", "FCS", "x", 5, COLOR_X_HEADER, "FCS"),
    ("V", "RCS", "x", 5, COLOR_X_HEADER, "RCS"),
    ("W", "POLICE", "x", 6, COLOR_X_HEADER, "POLICE"),
    ("X", "نوع\nالاشتراك", "auto", 12, COLOR_HEADER_BG, None),
    # التواريخ
    ("Y", "تاريخ\nآخر دفعة", "input", 12, COLOR_PAY_BG, None),
    ("Z", "تاريخ\nالانتهاء", "auto", 12, COLOR_PAY_BG, None),
    # حالة الدفع (4 خيارات)
    ("AA", "مدفوع", "x", 6, COLOR_X_HEADER, "مدفوع"),
    ("AB", "لم\nيدفع", "x", 6, COLOR_X_HEADER, "لم يدفع"),
    ("AC", "تأمين\nفقط", "x", 6, COLOR_X_HEADER, "تأمين فقط"),
    ("AD", "اشتراك\n300", "x", 6, COLOR_X_HEADER, "اشتراك 300"),
    ("AE", "حالة\nالدفع", "auto", 12, COLOR_PAY_BG, None),
    # الأموال
    ("AF", "رسوم\nالاشتراك", "auto", 10, COLOR_MONEY_BG, None),
    ("AG", "مصاريف\nالتأمين", "auto", 10, COLOR_MONEY_BG, None),
    ("AH", "حقوق\nالمركب", "auto", 10, COLOR_MONEY_BG, None),
    ("AI", "المبلغ\nالإجمالي", "auto", 12, COLOR_TOTAL_BG, None),
    ("AJ", "حالة\nالتجديد", "auto", 22, COLOR_RENEW_BG, None),
    # أيام السباحة (4 خيارات)
    ("AK", "الأحد\nوالأربعاء", "x", 8, COLOR_X_HEADER, "الأحد والأربعاء"),
    ("AL", "الاثنين\nوالخميس", "x", 8, COLOR_X_HEADER, "الاثنين والخميس"),
    ("AM", "الثلاثاء\nوالجمعة", "x", 8, COLOR_X_HEADER, "الثلاثاء والجمعة"),
    ("AN", "كل\nالأيام", "x", 6, COLOR_X_HEADER, "كل الأيام"),
    ("AO", "أيام\nالسباحة", "auto", 16, COLOR_TIME_BG, None),
    # التوقيت (4 خيارات)
    ("AP", "09:00\n10:00", "x", 7, COLOR_X_HEADER, "09:00-10:00"),
    ("AQ", "10:00\n11:00", "x", 7, COLOR_X_HEADER, "10:00-11:00"),
    ("AR", "19:00\n20:00", "x", 7, COLOR_X_HEADER, "19:00-20:00"),
    ("AS", "20:00\n21:00", "x", 7, COLOR_X_HEADER, "20:00-21:00"),
    ("AT", "التوقيت", "auto", 12, COLOR_TIME_BG, None),
    ("AU", "عدد أشهر\nالاشتراك", "auto", 9, COLOR_MONTHS_BG, None),
]

# عدد صفوف البيانات التي سيتم ملؤها
DATA_ROWS = 200
FIRST_DATA_ROW = 3
LAST_DATA_ROW = FIRST_DATA_ROW + DATA_ROWS - 1  # 202

# ───────────────── مساعدات ─────────────────

def col_idx(letter):
    """تحويل حرف العمود إلى رقم"""
    from openpyxl.utils import column_index_from_string
    return column_index_from_string(letter)

def make_font(size=10, bold=False, color="FF1F2937", name=FONT_NAME):
    return Font(name=name, size=size, bold=bold, color=color)

def make_fill(color):
    return PatternFill(patternType="solid", fgColor=color)

def make_align(h="center", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap, readingOrder=2)

# ───────────────── بناء ورقة بيانات ─────────────────

def build_data_sheet(wb):
    """بناء ورقة بيانات بالهيكل الجديد مع خانات X"""
    ws = wb.create_sheet("بيانات", 0)
    ws.sheet_view.rightToLeft = True
    ws.sheet_view.showGridLines = False

    # عناوين الأعمدة وعرضها
    for col_letter, header, kind, width, bg, x_opt in COLUMNS:
        ws.column_dimensions[col_letter].width = width

    # ─── صف العنوان (الصف 1) ───
    last_col_letter = COLUMNS[-1][0]
    ws.merge_cells(f"A1:{last_col_letter}1")
    title_cell = ws["A1"]
    title_cell.value = "   نظام إدارة اشتراكات  —  نادي RCS   "
    title_cell.font = Font(name=FONT_NAME, size=20, bold=True, color="FFFFFFFF")
    title_cell.fill = make_fill(COLOR_TITLE_BG)
    title_cell.alignment = make_align()
    ws.row_dimensions[1].height = 51.75

    # ─── صف العناوين (الصف 2) ───
    for col_letter, header, kind, width, bg, x_opt in COLUMNS:
        cell = ws[f"{col_letter}2"]
        cell.value = header
        cell.font = make_font(size=10, bold=True, color="FFFFFFFF")
        cell.fill = make_fill(bg)
        cell.alignment = make_align()
        cell.border = BORDER_ALL
    ws.row_dimensions[2].height = 42

    # ─── صفوف البيانات (3 إلى 202) ───
    # خريطة الأعمدة لاستخدامها في المعادلات
    col_map = {c[0]: c for c in COLUMNS}

    for r in range(FIRST_DATA_ROW, LAST_DATA_ROW + 1):
        ws.row_dimensions[r].height = 22

        for col_letter, header, kind, width, bg, x_opt in COLUMNS:
            cell = ws[f"{col_letter}{r}"]

            # تنسيق الخلية
            if kind == "x":
                # خانة علامة X
                cell.fill = make_fill(COLOR_X_BG)
                cell.font = make_font(size=12, bold=True, color="FFB45309")
                cell.alignment = make_align()
                cell.border = BORDER_ALL
            elif kind == "auto":
                if col_letter in ("AF", "AG", "AH"):
                    cell.font = Font(name=FONT_NAME_NUM, size=10, bold=True, color="FF0F766E")
                    cell.fill = make_fill(COLOR_AUTO_BG)
                elif col_letter == "AI":
                    cell.font = Font(name=FONT_NAME_NUM, size=11, bold=True, color="FFD97706")
                    cell.fill = make_fill("FFFFFBEB")
                elif col_letter == "AJ":
                    cell.font = make_font(size=10, bold=True, color="FF1F2937")
                    cell.fill = make_fill("FFEFF6FF")
                elif col_letter in ("A",):
                    cell.font = Font(name=FONT_NAME_NUM, size=10, bold=True, color="FF1B3A6B")
                    cell.fill = make_fill(COLOR_INPUT_BG)
                elif col_letter in ("G", "H", "Q", "X", "AE", "AO", "AT"):
                    cell.font = make_font(size=10, bold=True, color="FF0F766E")
                    cell.fill = make_fill(COLOR_AUTO_BG)
                elif col_letter == "Z":
                    cell.font = Font(name=FONT_NAME_NUM, size=10, color="FF1F2937")
                    cell.fill = make_fill(COLOR_INPUT_BG)
                elif col_letter == "AU":
                    cell.font = Font(name=FONT_NAME_NUM, size=10, bold=True, color="FFFFFFFF")
                    cell.fill = make_fill(COLOR_MONTHS_BG)
                else:
                    cell.font = make_font(size=10)
                    cell.fill = make_fill(COLOR_AUTO_BG)
                cell.alignment = make_align()
                cell.border = BORDER_ALL
            elif kind == "input":
                if col_letter in ("B", "C"):
                    cell.font = make_font(size=11, color="FF1F2937")
                    cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=False, readingOrder=2)
                elif col_letter in ("D", "Y"):
                    cell.font = Font(name=FONT_NAME_NUM, size=10, color="FF1F2937")
                    cell.alignment = make_align(wrap=False)
                cell.fill = make_fill(COLOR_INPUT_BG)
                cell.border = BORDER_ALL

        # ─── المعادلات ───
        # A: رقم الملف (تلقائي)
        ws[f"A{r}"] = f'=IF(B{r}="","","RCS "&TEXT(ROW()-2,"000"))'

        # G: الجنس (تلقائي من E,F)
        ws[f"G{r}"] = (
            f'=IF(B{r}="","",IF(E{r}<>"","ذكر",IF(F{r}<>"","أنثى","")))'
        )

        # H: العمر (تلقائي من D)
        ws[f"H{r}"] = f'=IF(D{r}="","",DATEDIF(D{r},TODAY(),"Y"))'

        # Q: فصيلة الدم (تلقائي من I-P)
        ws[f"Q{r}"] = (
            f'=IF(B{r}="","",IF(I{r}<>"","A+",IF(J{r}<>"","A-",'
            f'IF(K{r}<>"","B+",IF(L{r}<>"","B-",IF(M{r}<>"","O+",'
            f'IF(N{r}<>"","O-",IF(O{r}<>"","AB+",IF(P{r}<>"","AB-",""))))))))'
        )

        # X: نوع الاشتراك (تلقائي من R-W)
        ws[f"X{r}"] = (
            f'=IF(B{r}="","",IF(R{r}<>"","/",IF(S{r}<>"","OPOW",'
            f'IF(T{r}<>"","DJS",IF(U{r}<>"","FCS",IF(V{r}<>"","RCS",'
            f'IF(W{r}<>"","POLICE","")))))))'
        )

        # Z: تاريخ الانتهاء (تلقائي = Y + 30)
        ws[f"Z{r}"] = f'=IF(Y{r}="","",Y{r}+30)'

        # AE: حالة الدفع (تلقائي من AA-AD)
        ws[f"AE{r}"] = (
            f'=IF(B{r}="","",IF(AA{r}<>"","مدفوع",IF(AB{r}<>"","لم يدفع",'
            f'IF(AC{r}<>"","تأمين فقط",IF(AD{r}<>"","اشتراك 300","")))))'
        )

        # AF: رسوم الاشتراك
        # OPOW/DJS/POLICE/اشتراك 300 = 300
        # FCS/RCS/تأمين فقط = 0
        # عادي = 1300 (أقل من 14) أو 1500 (14+)
        ws[f"AF{r}"] = (
            f'=IF(OR(AE{r}="لم يدفع",B{r}=""),"",'
            f'IF(OR(AE{r}="تأمين فقط",X{r}="FCS",X{r}="RCS"),0,'
            f'IF(OR(AE{r}="اشتراك 300",X{r}="POLICE",X{r}="OPOW",X{r}="DJS"),300,'
            f'IF(H{r}<14,1300,1500))))'
        )

        # AG: مصاريف التأمين = 500 دج دائماً عند الدفع
        ws[f"AG{r}"] = f'=IF(OR(AE{r}="لم يدفع",B{r}=""),"",500)'

        # AH: حقوق المركب = 1000 (إلا للمستثناة)
        ws[f"AH{r}"] = (
            f'=IF(OR(AE{r}="لم يدفع",AE{r}="تأمين فقط",AE{r}="اشتراك 300",'
            f'X{r}="OPOW",X{r}="DJS",X{r}="FCS",X{r}="RCS",X{r}="POLICE",B{r}=""),"",1000)'
        )

        # AI: المبلغ الإجمالي = رسوم الاشتراك + مصاريف التأمين
        ws[f"AI{r}"] = (
            f'=IF(AE{r}="لم يدفع","",IF(AF{r}="","",AF{r}+AG{r}))'
        )

        # AJ: حالة التجديد
        ws[f"AJ{r}"] = (
            f'=IF(B{r}="","",IF(AE{r}="لم يدفع","🔒 مجمدة",'
            f'IF(Z{r}="","",IF(Z{r}<TODAY(),"⛔ منتهي - يتطلب تجديد",'
            f'IF(Z{r}-5<=TODAY(),"⚠️ قريب الانتهاء","✅ ساري")))))'
        )

        # AO: أيام السباحة (تلقائي من AK-AN)
        ws[f"AO{r}"] = (
            f'=IF(B{r}="","",IF(AK{r}<>"","الأحد والأربعاء",'
            f'IF(AL{r}<>"","الاثنين والخميس",IF(AM{r}<>"","الثلاثاء والجمعة",'
            f'IF(AN{r}<>"","كل الأيام","")))))'
        )

        # AT: التوقيت (تلقائي من AP-AS)
        ws[f"AT{r}"] = (
            f'=IF(B{r}="","",IF(AP{r}<>"","09:00-10:00",'
            f'IF(AQ{r}<>"","10:00-11:00",IF(AR{r}<>"","19:00-20:00",'
            f'IF(AS{r}<>"","20:00-21:00","")))))'
        )

        # AU: عدد أشهر الاشتراك
        ws[f"AU{r}"] = f'=IF(B{r}="","",COUNTA(التجديد!E{r}:E{r}))'

    # ─── التحقق من صحة البيانات (Data Validation) لخانات X ───
    # السماح فقط بـ X أو x أو ✓ أو •
    x_dv = DataValidation(
        type="list",
        formula1='"X,x,✓,•"',
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="علامة غير صالحة",
        error="ضع X أو • فقط",
        showInputMessage=True,
        promptTitle="خانة اختيار",
        prompt="ضع علامة X أو • في الخانة المختارة",
    )
    x_dv.ignoreBlank = True
    ws.add_data_validation(x_dv)

    # تطبيق التحقق على جميع خانات X
    x_columns = [c[0] for c in COLUMNS if c[2] == "x"]
    for col_letter in x_columns:
        rng = f"{col_letter}{FIRST_DATA_ROW}:{col_letter}{LAST_DATA_ROW}"
        x_dv.add(rng)

    # ─── تنسيق شرطي لإبراز خانات X المملوءة ───
    from openpyxl.styles.differential import DifferentialStyle
    from openpyxl.formatting.rule import Rule

    dxf = DifferentialStyle(
        font=Font(name=FONT_NAME, size=14, bold=True, color="FFDC2626"),
        fill=PatternFill(patternType="solid", fgColor="FFFEE2E2"),
    )
    rule = Rule(type="containsText", operator="containsText", text="X", dxf=dxf)
    rule.formula = ['NOT(ISERROR(SEARCH("X",CELL_ADDRESS)))']

    for col_letter in x_columns:
        rng = f"{col_letter}{FIRST_DATA_ROW}:{col_letter}{LAST_DATA_ROW}"
        # قاعدة بسيطة: إذا لم تكن الخلية فارغة
        ws.conditional_formatting.add(
            rng,
            CellIsRule(
                operator="notEqual",
                formula=['""'],
                fill=PatternFill(patternType="solid", fgColor="FFFEE2E2"),
                font=Font(name=FONT_NAME, size=14, bold=True, color="FFDC2626"),
            ),
        )

    # تنسيق الأعمدة الرقمية
    for r in range(FIRST_DATA_ROW, LAST_DATA_ROW + 1):
        ws[f"AF{r}"].number_format = '#,##0" دج"'
        ws[f"AG{r}"].number_format = '#,##0" دج"'
        ws[f"AH{r}"].number_format = '#,##0" دج"'
        ws[f"AI{r}"].number_format = '#,##0" دج"'
        ws[f"D{r}"].number_format = "DD/MM/YYYY"
        ws[f"Y{r}"].number_format = "DD/MM/YYYY"
        ws[f"Z{r}"].number_format = "DD/MM/YYYY"

    # تجميد الصفوف العلوية
    ws.freeze_panes = "B3"

    return ws


# ───────────────── بناء لوحة التحكم ─────────────────

def build_dashboard(wb):
    """بناء لوحة التحكم محدثة بالأعمدة الجديدة"""
    ws = wb.create_sheet("لوحة_التحكم")
    ws.sheet_view.rightToLeft = True
    ws.sheet_view.showGridLines = False

    # عرض الأعمدة
    widths = {"A": 2, "B": 24, "C": 18, "D": 2, "E": 22, "F": 12, "G": 2,
              "H": 18, "I": 10, "J": 2, "K": 22, "L": 18}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # صف العنوان
    ws.merge_cells("A1:L1")
    title = ws["A1"]
    title.value = "📊   لوحة التحكم  —  إحصائيات الاشتراكات  —  نادي RCS"
    title.font = Font(name=FONT_NAME, size=18, bold=True, color="FFFFFFFF")
    title.fill = make_fill(COLOR_TITLE_BG)
    title.alignment = make_align()
    ws.row_dimensions[1].height = 43.5

    # عناوين الأقسام (صف 3)
    sections = [
        ("B3", "💰  الإحصائيات المالية", COLOR_MONEY_BG),
        ("E3", "👥  إحصائيات المنخرطين", COLOR_HEADER_BG),
        ("H3", "📋  فئات الاشتراك", COLOR_BLOOD_BG),
        ("K3", "🔄  حالة التجديد", COLOR_RENEW_BG),
    ]
    for cell_ref, value, color in sections:
        ws[cell_ref] = value
        ws[cell_ref].font = make_font(size=11, bold=True, color="FFFFFFFF")
        ws[cell_ref].fill = make_fill(color)
        ws[cell_ref].alignment = make_align()
        ws[cell_ref].border = BORDER_ALL

    # صف 4-11: الإحصائيات المالية (B-C) و المنخرطين (E-F) وفئات الاشتراك (H-I) وحالة التجديد (K-L)
    # نطاق البيانات: AF3:AF200 (رسوم الاشتراك)، AG3:AG200 (التأمين)، AH3:AH200 (حقوق المركب)،
    # AI3:AI200 (المبلغ الإجمالي)، AE3:AE200 (حالة الدفع)، X3:X200 (نوع الاشتراك)،
    # B3:B200 (اللقب)، G3:G200 (الجنس)، H3:H200 (العمر)، AJ3:AJ200 (حالة التجديد)

    data_rows = f"3:{LAST_DATA_ROW}"

    # ─── صف 4 ───
    ws["B4"] = "إجمالي رسوم الاشتراك"
    ws["C4"] = f'=SUMIFS(بيانات!AF3:AF{LAST_DATA_ROW},بيانات!AE3:AE{LAST_DATA_ROW},"<>لم يدفع",بيانات!B3:B{LAST_DATA_ROW},"<>")'

    ws["E4"] = "إجمالي المنخرطين"
    ws["F4"] = f'=COUNTA(بيانات!B3:B{LAST_DATA_ROW})'

    ws["H4"] = "اشتراك عادي (/)"
    ws["I4"] = f'=COUNTIF(بيانات!X3:X{LAST_DATA_ROW},"/")'

    ws["K4"] = "✅ سارية"
    ws["L4"] = f'=COUNTIF(بيانات!AJ3:AJ{LAST_DATA_ROW},"*ساري*")'

    # ─── صف 5 ───
    ws["B5"] = "إجمالي مصاريف التأمين"
    ws["C5"] = f'=SUMIFS(بيانات!AG3:AG{LAST_DATA_ROW},بيانات!AE3:AE{LAST_DATA_ROW},"<>لم يدفع",بيانات!B3:B{LAST_DATA_ROW},"<>")'

    ws["E5"] = "✅ مدفوع"
    ws["F5"] = f'=COUNTIF(بيانات!AE3:AE{LAST_DATA_ROW},"مدفوع")'

    ws["H5"] = "OPOW"
    ws["I5"] = f'=COUNTIF(بيانات!X3:X{LAST_DATA_ROW},"OPOW")'

    ws["K5"] = "⚠️ قريبة الانتهاء"
    ws["L5"] = f'=COUNTIF(بيانات!AJ3:AJ{LAST_DATA_ROW},"*قريب*")'

    # ─── صف 6 ───
    ws["B6"] = "حقوق المركب (مستثناة)"
    ws["C6"] = f'=SUMIFS(بيانات!AH3:AH{LAST_DATA_ROW},بيانات!AE3:AE{LAST_DATA_ROW},"<>لم يدفع",بيانات!B3:B{LAST_DATA_ROW},"<>")'

    ws["E6"] = "❌ لم يدفع"
    ws["F6"] = f'=COUNTIF(بيانات!AE3:AE{LAST_DATA_ROW},"لم يدفع")'

    ws["H6"] = "DJS"
    ws["I6"] = f'=COUNTIF(بيانات!X3:X{LAST_DATA_ROW},"DJS")'

    ws["K6"] = "⛔ منتهية"
    ws["L6"] = f'=COUNTIF(بيانات!AJ3:AJ{LAST_DATA_ROW},"*منتهي*")'

    # ─── صف 7 ───
    ws["B7"] = "الإيرادات (رسوم + تأمين)"
    ws["C7"] = "=C4+C5"

    ws["E7"] = "🔵 تأمين فقط"
    ws["F7"] = f'=COUNTIF(بيانات!AE3:AE{LAST_DATA_ROW},"تأمين فقط")'

    ws["H7"] = "FCS"
    ws["I7"] = f'=COUNTIF(بيانات!X3:X{LAST_DATA_ROW},"FCS")'

    ws["K7"] = "🔒 مجمدة"
    ws["L7"] = f'=COUNTIF(بيانات!AJ3:AJ{LAST_DATA_ROW},"*مجمدة*")'

    # ─── صف 8 ───
    ws["B8"] = "اشتراكات عادية (1300)"
    ws["C8"] = f'=SUMIFS(بيانات!AI3:AI{LAST_DATA_ROW},بيانات!AF3:AF{LAST_DATA_ROW},1300,بيانات!AE3:AE{LAST_DATA_ROW},"<>لم يدفع")'

    ws["E8"] = "🟠 اشتراك 300"
    ws["F8"] = f'=COUNTIF(بيانات!AE3:AE{LAST_DATA_ROW},"اشتراك 300")'

    ws["H8"] = "RCS"
    ws["I8"] = f'=COUNTIF(بيانات!X3:X{LAST_DATA_ROW},"RCS")'

    # ─── صف 9 ───
    ws["B9"] = "اشتراكات عادية (1500)"
    ws["C9"] = f'=SUMIFS(بيانات!AI3:AI{LAST_DATA_ROW},بيانات!AF3:AF{LAST_DATA_ROW},1500,بيانات!AE3:AE{LAST_DATA_ROW},"<>لم يدفع")'

    ws["E9"] = "كبار (14+)"
    ws["F9"] = f'=COUNTIFS(بيانات!H3:H{LAST_DATA_ROW},">=14",بيانات!B3:B{LAST_DATA_ROW},"<>")'

    ws["H9"] = "POLICE"
    ws["I9"] = f'=COUNTIF(بيانات!X3:X{LAST_DATA_ROW},"POLICE")'

    # ─── صف 10 ───
    ws["B10"] = "إيرادات OPOW/DJS"
    ws["C10"] = (
        f'=SUMIFS(بيانات!AI3:AI{LAST_DATA_ROW},بيانات!X3:X{LAST_DATA_ROW},"OPOW",بيانات!AE3:AE{LAST_DATA_ROW},"<>لم يدفع")'
        f'+SUMIFS(بيانات!AI3:AI{LAST_DATA_ROW},بيانات!X3:X{LAST_DATA_ROW},"DJS",بيانات!AE3:AE{LAST_DATA_ROW},"<>لم يدفع")'
    )

    ws["E10"] = "صغار (أقل من 14)"
    ws["F10"] = f'=COUNTIFS(بيانات!H3:H{LAST_DATA_ROW},"<14",بيانات!H3:H{LAST_DATA_ROW},"<>",بيانات!B3:B{LAST_DATA_ROW},"<>")'

    # ─── صف 11 ───
    ws["B11"] = "متوسط الدفعة"
    ws["C11"] = (
        f'=IFERROR(AVERAGEIFS(بيانات!AI3:AI{LAST_DATA_ROW},بيانات!AE3:AE{LAST_DATA_ROW},"<>لم يدفع",بيانات!B3:B{LAST_DATA_ROW},"<>"),0)'
    )

    ws["E11"] = "ذكور"
    ws["F11"] = f'=COUNTIF(بيانات!G3:G{LAST_DATA_ROW},"ذكر")'

    ws["E12"] = "إناث"
    ws["F12"] = f'=COUNTIF(بيانات!G3:G{LAST_DATA_ROW},"أنثى")'

    # ─── صف 15-23: التوقيت/الأيام/فصيلة الدم/تفصيل مالي ───
    ws["B15"] = "🕐  أفواج التوقيت"
    ws["E15"] = "🏊  أيام السباحة"
    ws["H15"] = "🩸  فصائل الدم"
    ws["K15"] = "💳  ملخص مالي تفصيلي"

    # التوقيت (AT column)
    ws["B16"] = "09:00-10:00"
    ws["C16"] = f'=COUNTIF(بيانات!AT3:AT{LAST_DATA_ROW},"09:00-10:00")'
    ws["B17"] = "10:00-11:00"
    ws["C17"] = f'=COUNTIF(بيانات!AT3:AT{LAST_DATA_ROW},"10:00-11:00")'
    ws["B18"] = "19:00-20:00"
    ws["C18"] = f'=COUNTIF(بيانات!AT3:AT{LAST_DATA_ROW},"19:00-20:00")'
    ws["B19"] = "20:00-21:00"
    ws["C19"] = f'=COUNTIF(بيانات!AT3:AT{LAST_DATA_ROW},"20:00-21:00")'

    # أيام السباحة (AO column)
    ws["E16"] = "الأحد والأربعاء"
    ws["F16"] = f'=COUNTIF(بيانات!AO3:AO{LAST_DATA_ROW},"الأحد والأربعاء")'
    ws["E17"] = "الاثنين والخميس"
    ws["F17"] = f'=COUNTIF(بيانات!AO3:AO{LAST_DATA_ROW},"الاثنين والخميس")'
    ws["E18"] = "الثلاثاء والجمعة"
    ws["F18"] = f'=COUNTIF(بيانات!AO3:AO{LAST_DATA_ROW},"الثلاثاء والجمعة")'
    ws["E19"] = "كل الأيام"
    ws["F19"] = f'=COUNTIF(بيانات!AO3:AO{LAST_DATA_ROW},"كل الأيام")'

    # فصائل الدم (Q column)
    blood_types = [("A+", 16), ("A-", 17), ("B+", 18), ("B-", 19), ("O+", 20), ("O-", 21), ("AB+", 22), ("AB-", 23)]
    for bt, row in blood_types:
        ws[f"H{row}"] = bt
        ws[f"I{row}"] = f'=COUNTIF(بيانات!Q3:Q{LAST_DATA_ROW},"{bt}")'

    # ملخص مالي تفصيلي
    ws["K16"] = "رسوم 300 دج × عدد"
    ws["L16"] = f'=COUNTIF(بيانات!AF3:AF{LAST_DATA_ROW},300)&" منخرط"'
    ws["K17"] = "مجموع رسوم 300"
    ws["L17"] = f'=SUMIF(بيانات!AF3:AF{LAST_DATA_ROW},300,بيانات!AF3:AF{LAST_DATA_ROW})'
    ws["K18"] = "رسوم 1300 دج × عدد"
    ws["L18"] = f'=COUNTIF(بيانات!AF3:AF{LAST_DATA_ROW},1300)&" منخرط"'
    ws["K19"] = "مجموع رسوم 1300"
    ws["L19"] = f'=SUMIF(بيانات!AF3:AF{LAST_DATA_ROW},1300,بيانات!AF3:AF{LAST_DATA_ROW})'
    ws["K20"] = "رسوم 1500 دج × عدد"
    ws["L20"] = f'=COUNTIF(بيانات!AF3:AF{LAST_DATA_ROW},1500)&" منخرط"'
    ws["K21"] = "مجموع رسوم 1500"
    ws["L21"] = f'=SUMIF(بيانات!AF3:AF{LAST_DATA_ROW},1500,بيانات!AF3:AF{LAST_DATA_ROW})'
    ws["K22"] = "تأمين مُحصَّل"
    ws["L22"] = f'=SUMIF(بيانات!AE3:AE{LAST_DATA_ROW},"<>لم يدفع",بيانات!AG3:AG{LAST_DATA_ROW})'
    ws["K23"] = "حقوق مركب مُحصَّلة"
    ws["L23"] = f'=SUMIFS(بيانات!AH3:AH{LAST_DATA_ROW},بيانات!AE3:AE{LAST_DATA_ROW},"<>لم يدفع",بيانات!B3:B{LAST_DATA_ROW},"<>")'

    # تنسيق الخلايا
    # عناوين الأقسام (صف 3 و 15)
    for r in [3, 15]:
        for c in ["B", "E", "H", "K"]:
            cell = ws[f"{c}{r}"]
            if cell.value:
                cell.font = make_font(size=11, bold=True, color="FFFFFFFF")
                cell.fill = make_fill({"B": COLOR_MONEY_BG, "E": COLOR_HEADER_BG,
                                       "H": COLOR_BLOOD_BG, "K": COLOR_RENEW_BG}[c])
                cell.alignment = make_align()
                cell.border = BORDER_ALL

    # عناوين العناصر (العمود B, E, H, K)
    for r in range(4, 24):
        for c in ["B", "E", "H", "K"]:
            cell = ws[f"{c}{r}"]
            if cell.value:
                cell.font = make_font(size=10, bold=True, color="FF1F2937")
                cell.fill = make_fill("FFF8FAFC")
                cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True, readingOrder=2)
                cell.border = BORDER_ALL

    # القيم (العمود C, F, I, L)
    for r in range(4, 24):
        for c in ["C", "F", "I", "L"]:
            cell = ws[f"{c}{r}"]
            if cell.value:
                cell.font = Font(name=FONT_NAME_NUM, size=11, bold=True, color="FF0F766E")
                cell.fill = make_fill(COLOR_AUTO_BG)
                cell.alignment = make_align(wrap=False)
                cell.border = BORDER_ALL

    # تنسيق الأرقام
    for r in [4, 5, 6, 7, 8, 9, 10, 11, 17, 19, 21, 22, 23]:
        ws[f"C{r}"].number_format = '#,##0" دج"'
    for r in [4, 5, 6, 7, 8, 9, 10, 11, 12, 16, 17, 18, 19, 20, 21, 22, 23]:
        if ws[f"F{r}"].value:
            ws[f"F{r}"].number_format = "0"
        if ws[f"I{r}"].value:
            ws[f"I{r}"].number_format = "0"
        if ws[f"L{r}"].value:
            ws[f"L{r}"].number_format = "0"

    # تذييل
    ws.merge_cells("A35:L35")
    footer = ws["A35"]
    footer.value = "✦  نادي RCS  |  منظومة إدارة الاشتراكات  |  جميع الحقوق محفوظة  ✦"
    footer.font = Font(name="Arial", size=9, color="FFFFFFFF")
    footer.fill = make_fill(COLOR_FOOTER_BG)
    footer.alignment = make_align()
    ws.row_dimensions[35].height = 24

    return ws


# ───────────────── بناء ورقة حقوق المركب ─────────────────

def build_harbour_sheet(wb):
    """ورقة حقوق المركب - محدثة بالأعمدة الجديدة"""
    ws = wb.create_sheet("حقوق_المركب")
    ws.sheet_view.rightToLeft = True
    ws.sheet_view.showGridLines = False

    # عرض الأعمدة
    widths = {"A": 12, "B": 22, "C": 20, "D": 14, "E": 14, "F": 18, "G": 14}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # صف العنوان
    ws.merge_cells("A1:G1")
    title = ws["A1"]
    title.value = "🏊  سجل حقوق المركب  —  نادي RCS"
    title.font = Font(name=FONT_NAME, size=18, bold=True, color="FFFFFFFF")
    title.fill = make_fill(COLOR_TITLE_BG)
    title.alignment = make_align()
    ws.row_dimensions[1].height = 43.5

    # العناوين
    headers = ["رقم الملف", "اللقب", "الاسم", "رسوم الاشتراك", "حقوق المركب", "المبلغ الإجمالي", "حالة الدفع"]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=i)
        cell.value = h
        cell.font = make_font(size=10, bold=True, color="FFFFFFFF")
        cell.fill = make_fill(COLOR_HEADER_BG)
        cell.alignment = make_align()
        cell.border = BORDER_ALL
    ws.row_dimensions[2].height = 36

    # البيانات - الآن الأعمدة الجديدة: A,B,C,AF,AH,AI,AE
    # AF = رسوم الاشتراك، AH = حقوق المركب، AI = المبلغ الإجمالي، AE = حالة الدفع
    for r in range(3, LAST_DATA_ROW + 1):
        ws.row_dimensions[r].height = 22
        ws[f"A{r}"] = f'=IF(بيانات!AH{r}=1000,بيانات!A{r},"")'
        ws[f"B{r}"] = f'=IF(بيانات!AH{r}=1000,بيانات!B{r},"")'
        ws[f"C{r}"] = f'=IF(بيانات!AH{r}=1000,بيانات!C{r},"")'
        ws[f"D{r}"] = f'=IF(بيانات!AH{r}=1000,بيانات!AF{r},"")'
        ws[f"E{r}"] = f'=IF(بيانات!AH{r}=1000,بيانات!AH{r},"")'
        ws[f"F{r}"] = f'=IF(بيانات!AH{r}=1000,بيانات!AI{r},"")'
        ws[f"G{r}"] = f'=IF(بيانات!AH{r}=1000,بيانات!AE{r},"")'

        # تنسيق
        for col_letter in ["A", "B", "C", "D", "E", "F", "G"]:
            cell = ws[f"{col_letter}{r}"]
            cell.font = make_font(size=10) if col_letter in ["B", "C"] else Font(name=FONT_NAME_NUM, size=10)
            cell.fill = make_fill(COLOR_INPUT_BG)
            cell.alignment = make_align(wrap=False) if col_letter not in ["B", "C"] else Alignment(
                horizontal="right", vertical="center", readingOrder=2)
            cell.border = BORDER_ALL
        ws[f"D{r}"].number_format = '#,##0" دج"'
        ws[f"E{r}"].number_format = '#,##0" دج"'
        ws[f"F{r}"].number_format = '#,##0" دج"'

    ws.freeze_panes = "A3"
    return ws


# ───────────────── بناء ورقة التجديد ─────────────────

def build_renewal_sheet(wb):
    """ورقة التجديد - محدثة بالأعمدة الجديدة"""
    ws = wb.create_sheet("التجديد")
    ws.sheet_view.rightToLeft = True
    ws.sheet_view.showGridLines = False

    widths = {"A": 12, "B": 22, "C": 20, "D": 18, "E": 16, "F": 18, "G": 14, "H": 14, "I": 22}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # عنوان
    ws.merge_cells("A1:I1")
    title = ws["A1"]
    title.value = "🔄  سجل تجديد الاشتراكات  —  نادي RCS"
    title.font = Font(name=FONT_NAME, size=18, bold=True, color="FFFFFFFF")
    title.fill = make_fill(COLOR_TITLE_BG)
    title.alignment = make_align()
    ws.row_dimensions[1].height = 43.5

    headers = ["رقم الملف", "اللقب", "الاسم", "نهاية الاشتراك السابق", "تاريخ التجديد",
               "نهاية الشهر الجديد", "مبلغ التجديد", "حالة الدفع", "حالة التجديد"]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=i)
        cell.value = h
        cell.font = make_font(size=10, bold=True, color="FFFFFFFF")
        cell.fill = make_fill(COLOR_HEADER_BG)
        cell.alignment = make_align()
        cell.border = BORDER_ALL
    ws.row_dimensions[2].height = 36

    # البيانات - الأعمدة الجديدة: A,B,C,Z(تاريخ الانتهاء),K→AE,P→AJ
    for r in range(3, LAST_DATA_ROW + 1):
        ws.row_dimensions[r].height = 22
        ws[f"A{r}"] = f'=IF(بيانات!B{r}<>"",بيانات!A{r},"")'
        ws[f"B{r}"] = f'=IF(بيانات!B{r}<>"",بيانات!B{r},"")'
        ws[f"C{r}"] = f'=IF(بيانات!B{r}<>"",بيانات!C{r},"")'
        ws[f"D{r}"] = f'=IF(بيانات!B{r}<>"",بيانات!Z{r},"")'  # Z = تاريخ الانتهاء
        # E: تاريخ التجديد (إدخال يدوي)
        ws[f"F{r}"] = f'=IF(E{r}="","",E{r}+30)'
        ws[f"H{r}"] = f'=IF(بيانات!B{r}<>"",بيانات!AE{r},"")'  # AE = حالة الدفع
        ws[f"I{r}"] = f'=IF(بيانات!B{r}<>"",بيانات!AJ{r},"")'  # AJ = حالة التجديد

        for col_letter in ["A", "B", "C", "D", "E", "F", "G", "H", "I"]:
            cell = ws[f"{col_letter}{r}"]
            if col_letter in ["B", "C", "I"]:
                cell.font = make_font(size=10)
                cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True, readingOrder=2)
            else:
                cell.font = Font(name=FONT_NAME_NUM, size=10)
                cell.alignment = make_align(wrap=False)
            cell.fill = make_fill(COLOR_INPUT_BG)
            cell.border = BORDER_ALL
        ws[f"D{r}"].number_format = "DD/MM/YYYY"
        ws[f"E{r}"].number_format = "DD/MM/YYYY"
        ws[f"F{r}"].number_format = "DD/MM/YYYY"
        ws[f"G{r}"].number_format = '#,##0" دج"'

    ws.freeze_panes = "A3"
    return ws


# ───────────────── بناء ورقة فئات الأعمار ─────────────────

def build_age_categories_sheet(wb):
    """ورقة تصنيف الأعمار - محدثة بالأعمدة الجديدة (G=الجنس، H=العمر)"""
    ws = wb.create_sheet("فئات_الأعمار")
    ws.sheet_view.rightToLeft = True
    ws.sheet_view.showGridLines = False

    # عرض الأعمدة
    for col_letter in "ABCDEFGHIJKLMNOPQRSTUV":
        ws.column_dimensions[col_letter].width = 14
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 8

    # عنوان
    ws.merge_cells("A1:V1")
    title = ws["A1"]
    title.value = "🏷️   تصنيف المنخرطين حسب الفئة العمرية والجنس  —  نادي RCS"
    title.font = Font(name=FONT_NAME, size=18, bold=True, color="FFFFFFFF")
    title.fill = make_fill(COLOR_TITLE_BG)
    title.alignment = make_align()
    ws.row_dimensions[1].height = 43.5

    # عناوين الفئات (صف 3)
    categories = [
        ("A3", "فئة الذكور أقل من 13 سنة", COLOR_HEADER_BG),
        ("E3", "فئة الإناث أقل من 13 سنة", "FFBE185D"),
        ("I3", "فئة الذكور 13 سنة فما فوق", COLOR_HEADER_BG),
        ("M3", "فئة الإناث 13 سنة فما فوق", "FFBE185D"),
    ]
    for cell_ref, value, color in categories:
        ws[cell_ref] = value
        ws[cell_ref].font = make_font(size=11, bold=True, color="FFFFFFFF")
        ws[cell_ref].fill = make_fill(color)
        ws[cell_ref].alignment = make_align()
        ws[cell_ref].border = BORDER_ALL
        # دمج الخلايا
        col = cell_ref[0]
        next_col = chr(ord(col) + 2)
        ws.merge_cells(f"{cell_ref}:{next_col}3")

    # صف 4: العدادات
    # الجنس الآن في العمود G، والعمر في العمود H
    ws["A4"] = '="العدد: " & COUNTIFS(بيانات!$G$3:$G$200,"ذكر",بيانات!$H$3:$H$200,"<13")'
    ws["E4"] = '="العدد: " & COUNTIFS(بيانات!$G$3:$G$200,"أنثى",بيانات!$H$3:$H$200,"<13")'
    ws["I4"] = '="العدد: " & COUNTIFS(بيانات!$G$3:$G$200,"ذكر",بيانات!$H$3:$H$200,">=13")'
    ws["M4"] = '="العدد: " & COUNTIFS(بيانات!$G$3:$G$200,"أنثى",بيانات!$H$3:$H$200,">=13")'
    for cell_ref in ["A4", "E4", "I4", "M4"]:
        cell = ws[cell_ref]
        cell.font = make_font(size=10, bold=True, color="FF1F2937")
        cell.fill = make_fill("FFFEF3C7")
        cell.alignment = make_align()
        cell.border = BORDER_ALL
        col = cell_ref[0]
        next_col = chr(ord(col) + 2)
        ws.merge_cells(f"{cell_ref}:{next_col}4")

    # صف 5: عناوين الأعمدة لكل فئة
    sub_headers = [
        ("A5", "رقم الملف", "B5", "اللقب والاسم", "C5", "العمر"),
        ("E5", "رقم الملف", "F5", "اللقب والاسم", "G5", "العمر"),
        ("I5", "رقم الملف", "J5", "اللقب والاسم", "K5", "العمر"),
        ("M5", "رقم الملف", "N5", "اللقب والاسم", "O5", "العمر"),
    ]
    for a, av, b, bv, c, cv in sub_headers:
        ws[a] = av
        ws[b] = bv
        ws[c] = cv
        for ref in [a, b, c]:
            cell = ws[ref]
            cell.font = make_font(size=10, bold=True, color="FFFFFFFF")
            cell.fill = make_fill(COLOR_HEADER_BG)
            cell.alignment = make_align()
            cell.border = BORDER_ALL

    # أعمدة المساعدة S-V (خفية)
    # S: ذكر <13, T: أنثى <13, U: ذكر >=13, V: أنثى >=13
    # الجنس في G، العمر في H
    helper_headers = [("S2", "ذكر<13"), ("T2", "أنثى<13"), ("U2", "ذكر≥13"), ("V2", "أنثى≥13")]
    for ref, val in helper_headers:
        ws[ref] = val
        ws[ref].font = make_font(size=8, color="FF6B7280")

    # بيانات التصنيف (صفوف 3-200 من بيانات)
    for r in range(3, LAST_DATA_ROW + 1):
        ws[f"S{r}"] = f'=IF(AND(بيانات!$B{r}<>"",بيانات!$G{r}="ذكر",بيانات!$H{r}<13),COUNTA($S$3:S{r}),"")'
        ws[f"T{r}"] = f'=IF(AND(بيانات!$B{r}<>"",بيانات!$G{r}="أنثى",بيانات!$H{r}<13),COUNTA($T$3:T{r}),"")'
        ws[f"U{r}"] = f'=IF(AND(بيانات!$B{r}<>"",بيانات!$G{r}="ذكر",بيانات!$H{r}>=13),COUNTA($U$3:U{r}),"")'
        ws[f"V{r}"] = f'=IF(AND(بيانات!$B{r}<>"",بيانات!$G{r}="أنثى",بيانات!$H{r}>=13),COUNTA($V$3:V{r}),"")'

    # إخفاء أعمدة المساعد
    for col_letter in ["S", "T", "U", "V"]:
        ws.column_dimensions[col_letter].hidden = True

    # صفوف العرض (6 إلى 105)
    for display_r in range(6, 106):
        data_r = display_r - 3  # offset
        # فئة الذكور <13
        ws[f"A{display_r}"] = f'=IFERROR(INDEX(بيانات!$A$3:$A${LAST_DATA_ROW},MATCH({data_r-2},$S$3:$S${LAST_DATA_ROW},0)),"")'
        ws[f"B{display_r}"] = f'=IFERROR(INDEX(بيانات!$B$3:$B${LAST_DATA_ROW},MATCH({data_r-2},$S$3:$S${LAST_DATA_ROW},0))&" "&INDEX(بيانات!$C$3:$C${LAST_DATA_ROW},MATCH({data_r-2},$S$3:$S${LAST_DATA_ROW},0)),"")'
        ws[f"C{display_r}"] = f'=IFERROR(INDEX(بيانات!$H$3:$H${LAST_DATA_ROW},MATCH({data_r-2},$S$3:$S${LAST_DATA_ROW},0)),"")'

        # فئة الإناث <13
        ws[f"E{display_r}"] = f'=IFERROR(INDEX(بيانات!$A$3:$A${LAST_DATA_ROW},MATCH({data_r-2},$T$3:$T${LAST_DATA_ROW},0)),"")'
        ws[f"F{display_r}"] = f'=IFERROR(INDEX(بيانات!$B$3:$B${LAST_DATA_ROW},MATCH({data_r-2},$T$3:$T${LAST_DATA_ROW},0))&" "&INDEX(بيانات!$C$3:$C${LAST_DATA_ROW},MATCH({data_r-2},$T$3:$T${LAST_DATA_ROW},0)),"")'
        ws[f"G{display_r}"] = f'=IFERROR(INDEX(بيانات!$H$3:$H${LAST_DATA_ROW},MATCH({data_r-2},$T$3:$T${LAST_DATA_ROW},0)),"")'

        # فئة الذكور >=13
        ws[f"I{display_r}"] = f'=IFERROR(INDEX(بيانات!$A$3:$A${LAST_DATA_ROW},MATCH({data_r-2},$U$3:$U${LAST_DATA_ROW},0)),"")'
        ws[f"J{display_r}"] = f'=IFERROR(INDEX(بيانات!$B$3:$B${LAST_DATA_ROW},MATCH({data_r-2},$U$3:$U${LAST_DATA_ROW},0))&" "&INDEX(بيانات!$C$3:$C${LAST_DATA_ROW},MATCH({data_r-2},$U$3:$U${LAST_DATA_ROW},0)),"")'
        ws[f"K{display_r}"] = f'=IFERROR(INDEX(بيانات!$H$3:$H${LAST_DATA_ROW},MATCH({data_r-2},$U$3:$U${LAST_DATA_ROW},0)),"")'

        # فئة الإناث >=13
        ws[f"M{display_r}"] = f'=IFERROR(INDEX(بيانات!$A$3:$A${LAST_DATA_ROW},MATCH({data_r-2},$V$3:$V${LAST_DATA_ROW},0)),"")'
        ws[f"N{display_r}"] = f'=IFERROR(INDEX(بيانات!$B$3:$B${LAST_DATA_ROW},MATCH({data_r-2},$V$3:$V${LAST_DATA_ROW},0))&" "&INDEX(بيانات!$C$3:$C${LAST_DATA_ROW},MATCH({data_r-2},$V$3:$V${LAST_DATA_ROW},0)),"")'
        ws[f"O{display_r}"] = f'=IFERROR(INDEX(بيانات!$H$3:$H${LAST_DATA_ROW},MATCH({data_r-2},$V$3:$V${LAST_DATA_ROW},0)),"")'

        # تنسيق صفوف العرض
        for col_letter in ["A", "B", "C", "E", "F", "G", "I", "J", "K", "M", "N", "O"]:
            cell = ws[f"{col_letter}{display_r}"]
            cell.font = make_font(size=10)
            cell.alignment = make_align(wrap=False)
            cell.fill = make_fill(COLOR_INPUT_BG)
            cell.border = BORDER_ALL

    # تعديل عرض الأعمدة
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["F"].width = 28
    ws.column_dimensions["J"].width = 28
    ws.column_dimensions["N"].width = 28
    ws.column_dimensions["D"].width = 3
    ws.column_dimensions["H"].width = 3
    ws.column_dimensions["L"].width = 3
    ws.column_dimensions["P"].width = 3

    ws.freeze_panes = "A6"
    return ws


# ───────────────── البرنامج الرئيسي ─────────────────

def main():
    print("🚀 بدء بناء منظومة RCS v3...")
    wb = Workbook()
    # حذف الورقة الافتراضية
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    print("📋 بناء ورقة بيانات...")
    build_data_sheet(wb)

    print("📊 بناء لوحة التحكم...")
    build_dashboard(wb)

    print("🏊 بناء ورقة حقوق المركب...")
    build_harbour_sheet(wb)

    print("🔄 بناء ورقة التجديد...")
    build_renewal_sheet(wb)

    print("🏷️ بناء ورقة فئات الأعمار...")
    build_age_categories_sheet(wb)

    # خصائص المصنف
    wb.properties.creator = "Z.ai"
    wb.properties.title = "منظومة اشتراكات RCS v3"
    wb.properties.description = "منظومة إدارة اشتراكات نادي RCS مع خانات اختيار"

    # حفظ
    os.makedirs(os.path.dirname(DST), exist_ok=True)
    wb.save(DST)
    print(f"✅ تم الحفظ: {DST}")
    print(f"📦 حجم الملف: {os.path.getsize(DST) / 1024:.1f} KB")


if __name__ == "__main__":
    main()
